#Con esta funcion vamos cargando a los empleados de la empresa
def empleados(cantidad):
    lista = []
    cont_empleados = 0
    if cont_empleados < cantidad:
        import datetime
        legajo = int(input('Ingrese numero de legajo del trabajador: '))
        nombre = input('Introduzca el nombre del trabajador: ')
        apellido = input('Introduzca el apellido del trabajador: ')
        edad = int(input('Introduzca la edad del trabajador: '))
        # Introduce la fecha de nacimiento en formato 'YYYY-MM-DD'
        fecha_de_ingreso = input("Introduzca la fecha de ingreso del empleado (DD-MM-YYYY): ")

        puesto = input('Introduzca el puesto que ocupa del trabajador: ')
        mail = input('Ingrese el mail del trabajador: ')
        empleado = [legajo, nombre, apellido, edad, mail, fecha_de_ingreso, puesto]
        lista.append(empleado)
        cont_empleados = cont_empleados + 1
    return lista


def antiguedad(lista): 
    name = input('Ingrese el nombre de la persona a consultar: ')
    surname = input('Ingrese el apellido de la persona a consultar: ')
    for x in lista:
        if name == lista[1] and surname == lista[2]:  
            import datetime
            fecha_de_ingreso = datetime.datetime.strptime(lista[5], '%d-%m-%Y')
            anios = datetime.datetime.now().year - fecha_de_ingreso
    print('La antiguedad de',surname,name,'es:',anios)  




def recategorizacion(lista):
    import datetime
    for y in lista:
        name = input('Ingrese nombre del empleado al que desea recategorizar o saber su categorizacion: ')
        surname = input('Ingrese apellido del empleado al que desea recategorizar o saber su categorizacion: ')
        if y[1] == name and y[2] == surname:
            fecha_de_ingreso = datetime.datetime.strptime(lista[5], '%d-%m-%Y')
            anios = datetime.datetime.now().year - fecha_de_ingreso
            anios_previos = int(input('Ingrese los años que tiene en el puesto antes de entrar a la empresa: '))
            antiguedad_total = anios + anios_previos
            if 0 <= antiguedad_total <= 1:
                print('Este empleado estaria en la categoria Traniee')
            elif 1 < antiguedad_total <= 3:
                print('Este empleado estaria en la categoria Junior')
            elif 3 < antiguedad_total <= 6:
                print('Este empleado estaria en la categoria Semi Senior')
            elif 6 < antiguedad_total:
                print('Este empleado estaria en la categoria Senior') 
            else: 
                print('Opcion incorrecta')
    


print('Menú de opciones de consulta para recursos humanos')
print('a) Permite cargar la lista de empleados')
print('b) Imprime la lista de empleados de la organización.')
print('c) Imprime la antiguedad del trabajador que desea saber')
print('d) Imprime si el trabajador requiere una recategorizacion de trainee, jr, semi senior o senior dentro de la organización.')
print('e) Enviar mail al empleado (opcion no valida aun)')
print('f) Desea salir del programa')

opcion = input('Ingrese una opción: ')
while opcion != 'f':
    if opcion == 'a':
        cant = int(input('Ingrese la cantidad de empleados: '))
        x = empleados(cant)
    elif opcion == 'b':
        print(empleados)
    elif opcion == 'c':
        print(antiguedad(empleados))
    elif opcion == 'd':
        recategorizacion(empleados)
    elif opcion == 'e':
        """ import smtplib 
        from email.message import EmailMessage 
        email_subject = "Email test from Python" 
        sender_email_address = "your_email@gmail.com" 
        receiver_email_address = "receiver_email@address.com" 
        email_smtp = "smtp.gmail.com" 
        email_password = "your_email_password" 

        # Create an email message object 
        message = EmailMessage() 

        # Configure email headers 
        message['Subject'] = email_subject 
        message['From'] = sender_email_address 
        message['To'] = receiver_email_address 

        # Set email body text 
        message.set_content("Hello from Python!") 

        # Set smtp server and port 
        server = smtplib.SMTP(email_smtp, '587') 

        # Identify this client to the SMTP server 
        server.ehlo() 

        # Secure the SMTP connection 
        server.starttls() 

        # Login to email account 
        server.login(sender_email_address, email_password) 

        # Send email 
        server.send_message(message) 

        # Close connection to server 
        server.quit()"""    
    else: 
        print('La opción seleccionada es incorrecto') 
    opcion = input('Ingrese una opción entre a, b, c, d, e o f para finalizar: ')


print('Gracias por utilizar nuestros servicios. Que tenga un lindo día')




""" from openpyxl import Workbook
from openpyxl import load_workbook
 """

#Creacion del excel
""" wb = Workbook ()
wb.save('testing.xlsx') """

""" wb2 = load_workbook('testing.xlsx')
 """
#Creacion de las hojas de excel
""" ws1 = wb2.create_sheet('hoja1')
ws2 = wb2.create_sheet('hoja2',0)
ws3 = wb2.create_sheet
wb2.save('testing.xlsx') """

#Como cambiarles el nombre a la hoja excel a gusto y conveniencia
""" ws = wb2['hoja2']
ws.title = 'Empleados'
wb2.save('testing.xlsx') """

#Accediendo a los datos (cambiar el numero de la celda para poder ir accediendo a los resultados de los datos)
""" ws = wb2['Empleados']
ws['E5'] = 'nombre'
 """
#Saber el valor de la celda
""" x = ws['E6']
print (x.value)
 """

#Accediendo a un grupo de celdas
""" cell_range = ws['A1':'G10'] """
""" print(cell_range)
print(cell_range[0][0].value) """

""" nombres = ws['A2':'A10']
contador = -1
for x in nombres:
    contador += 1
    if (x[0].value == 'Pascual'):
        print('El mail es '+ x[0].value+str(cell_range[contador][3].value))

wb2.save('testing.xlsx') """



""" #import openpyxl

# Cargar datos desde un archivo de Excel
workbook = openpyxl.load_workbook('archivo.xlsx')
sheet = workbook.active
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)
print(data)

# Guardar datos en un archivo de Excel
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
data = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
for row in data:
    new_sheet.append(row)
new_workbook.save('nuevo_archivo.xlsx')


# Cargar datos desde un archivo de texto
with open('archivo.txt', 'r') as file:
    data = file.read()
    print(data)

# Guardar datos en un archivo de texto
data = 'Esto es un ejemplo de texto que se guardará en el archivo.'
with open('nuevo_archivo.txt', 'w') as file:
    file.write(data)

 """


""" import smtplib 
from email.message import EmailMessage 

email_subject = "Email test from Python" 
sender_email_address = "your_email@gmail.com" 
receiver_email_address = "receiver_email@address.com" 
email_smtp = "smtp.gmail.com" 
email_password = "your_email_password" 

# Create an email message object 
message = EmailMessage() 

# Configure email headers 
message['Subject'] = email_subject 
message['From'] = sender_email_address 
message['To'] = receiver_email_address 

# Set email body text 
message.set_content("Hello from Python!") 

# Set smtp server and port 
server = smtplib.SMTP(email_smtp, '587') 

# Identify this client to the SMTP server 
server.ehlo() 

# Secure the SMTP connection 
server.starttls() 

# Login to email account 
server.login(sender_email_address, email_password) 

# Send email 
server.send_message(message) 

# Close connection to server 
server.quit() 






import datetime

# Introduce la fecha de nacimiento en formato 'YYYY-MM-DD'
fecha_de_ingreso = input("Introduzca la fecha de ingreso del empleado (DD-MM-YYYY): ")
# Calcula la edad a partir de la fecha de nacimiento
fecha_de_ingreso = datetime.datetime.strptime(fecha_de_ingreso, '%d-%m-%Y')
edad = datetime.datetime.now().year - fecha_de_ingreso.year
# Imprime la edad calculada

print("La antigüedad de:", edad)






import datetime

# Introduce la fecha de nacimiento en formato 'YYYY-MM-DD'
fecha_de_ingreso = input("Introduzca la fecha de ingreso del empleado (DD-MM-YYYY): ")
# Calcula la edad a partir de la fecha de nacimiento
fecha_de_ingreso = datetime.datetime.strptime(fecha_de_ingreso, '%d-%m-%Y')
edad = datetime.datetime.now().year - fecha_de_ingreso.year
# Imprime la edad calculada

print("La antigüedad de:", edad) """




        
#Calcular la antiguedad y que el programa envie una alerta de que le llegara una potencial recategorizacion en su puesto